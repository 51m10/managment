import flet as ft
import openpyxl
from pypdf import PdfReader
import shutil
import os

def main(page: ft.Page):
    page.title = "Factory Doc App"
    page.vertical_alignment = ft.MainAxisAlignment.CENTER
    page.horizontal_alignment = ft.CrossAxisAlignment.CENTER
    page.padding = 20
    page.rtl = True

    # ایجاد پوشه امن داخلی برای ذخیره فایل انتخاب شده
    docs_dir = page.get_app_storage_dir() if hasattr(page, 'get_app_storage_dir') else "."
    if not os.path.exists(docs_dir):
        os.makedirs(docs_dir, exist_ok=True)

    status_text = ft.Text("لطفاً با دکمه زیر فایل شرکت را انتخاب کنید.", size=12, weight=ft.FontWeight.BOLD)

    # اندازه فونت متن نتایج (قابل تغییر با دکمه‌های بزرگ‌نمایی/کوچک‌نمایی)
    result_font_size = {"value": 14}

    company_dropdown = ft.Dropdown(
        label="انتخاب فایل اسناد",
        hint_text="فایلی انتخاب نشده است",
        expand=True,
        border_radius=10
    )

    # کانتینر نمایش نتیجه جستجو (متن استخراج شده از صفحات PDF)
    results_view = ft.Column(
        controls=[],
        scroll=ft.ScrollMode.AUTO,
        expand=True,
    )
    # نگهداری رفرنس به کنترل‌های متنیِ نتایج، برای اعمال بزرگ‌نمایی/کوچک‌نمایی
    result_text_controls = []

    def update_dropdown():
        files = []
        if os.path.exists(docs_dir):
            files = [f for f in os.listdir(docs_dir) if f.endswith(('.xlsx', '.pdf'))]
        company_dropdown.options = [ft.dropdown.Option(f) for f in files]
        if files:
            status_text.value = f"تعداد {len(files)} فایل آماده است."
        else:
            status_text.value = "فایلی انتخاب نشده است."
        page.update()

    # تعریف فایل‌پیکر و اضافه کردن به overlay صفحه (بخش حیاتی برای اندروید)
    file_picker = ft.FilePicker()
    page.overlay.append(file_picker)

    def on_dialog_result(e: ft.FilePickerResultEvent):
        if e.files and len(e.files) > 0:
            selected_file = e.files[0]
            source_path = selected_file.path
            file_name = selected_file.name

            dest_path = os.path.join(docs_dir, file_name)
            try:
                # کپی کردن فایل انتخاب شده به پوشه داخلی اپ
                if source_path != dest_path:
                    shutil.copy(source_path, dest_path)

                update_dropdown()
                company_dropdown.value = file_name
                status_text.value = f"فایل '{file_name}' با موفقیت اضافه شد."
            except Exception:
                status_text.value = "خطا در بارگذاری فایل انتخاب شده."
            page.update()

    file_picker.on_result = on_dialog_result

    def process_file(e):
        selected_file = company_dropdown.value
        if not selected_file:
            status_text.value = "لطفاً ابتدا یک فایل را انتخاب کنید."
            page.update()
            return

        file_path = os.path.join(docs_dir, selected_file)
        try:
            if file_path.endswith('.xlsx'):
                wb = openpyxl.load_workbook(file_path, read_only=True)
                status_text.value = f"فایل اکسل بار شد. شیت‌ها: {', '.join(wb.sheetnames)}"
            elif file_path.endswith('.pdf'):
                reader = PdfReader(file_path)
                status_text.value = f"فایل PDF بار شد. تعداد صفحات: {len(reader.pages)}"
        except Exception:
            status_text.value = "خطا در خواندن محتوای فایل."
        page.update()

    # دکمه فراخوانی فایل‌پیکر با فیلتر روی فایل‌های اکسل و PDF
    pick_button = ft.TextButton(
        content=ft.Row([ft.Icon(ft.Icons.FOLDER_OPEN), ft.Text("انتخاب فایل از گوشی")], spacing=5),
        on_click=lambda _: file_picker.pick_files(
            allow_multiple=False,
            allowed_extensions=["xlsx", "pdf"]
        )
    )

    process_button = ft.TextButton(
        content=ft.Row([ft.Icon(ft.Icons.CHECK), ft.Text("پردازش فایل")], spacing=5),
        on_click=process_file
    )

    search_input = ft.TextField(
        hint_text="جستجو در اسناد...",
        prefix_icon=ft.Icons.SEARCH,
        border_radius=10,
        expand=True
    )

    def find_excel_file():
        """اولین فایل xlsx موجود در پوشه اسناد را برمی‌گرداند."""
        if not os.path.exists(docs_dir):
            return None
        for f in os.listdir(docs_dir):
            if f.endswith('.xlsx'):
                return os.path.join(docs_dir, f)
        return None

    def find_matching_row(excel_path, query):
        """
        در فایل اکسل به دنبال ردیفی می‌گردد که ستون 'اسم دستگاه' شامل عبارت جستجو باشد.
        خروجی: (pdf_filename, start_page, end_page, device_name) یا None
        """
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        ws = wb[wb.sheetnames[0]]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None

        header = [str(h).strip() if h else "" for h in rows[0]]

        # پیدا کردن ایندکس ستون‌ها بر اساس نام هدر (مقاوم در برابر جابجایی ستون‌ها)
        def col_index(possible_names):
            for name in possible_names:
                for i, h in enumerate(header):
                    if name in h:
                        return i
            return None

        link_idx = col_index(["لینک", "Hyperlink", "link"])
        name_idx = col_index(["اسم دستگاه", "دستگاه", "نام"])
        start_idx = col_index(["شروع"])
        end_idx = col_index(["پایان"])

        if None in (link_idx, name_idx, start_idx, end_idx):
            return None

        query_norm = query.strip().lower()

        for row in rows[1:]:
            if row is None or len(row) <= max(link_idx, name_idx, start_idx, end_idx):
                continue
            device_name = str(row[name_idx]).strip() if row[name_idx] else ""
            if query_norm in device_name.lower():
                link_value = str(row[link_idx]).strip() if row[link_idx] else ""
                # لینک به فرم "1.pdf#page=1" است؛ فقط نام فایل را جدا می‌کنیم
                pdf_filename = link_value.split('#')[0].strip()
                start_page = int(row[start_idx])
                end_page = int(row[end_idx])
                return pdf_filename, start_page, end_page, device_name

        return None

    def split_into_paragraphs(text):
        """متن پاکسازی‌شده را بر اساس خط خالی به پاراگراف‌های جداگانه تقسیم می‌کند."""
        parts = [p.strip() for p in text.split("\n\n")]
        return [p for p in parts if p]

    def build_highlighted_spans(text, query, base_size):
        """
        بخش‌هایی از متن که با عبارت جستجو مطابقت دارند را پررنگ و رنگی
        نمایش می‌دهد تا پیدا کردنشان در متن راحت‌تر باشد.
        """
        spans = []
        if not query:
            spans.append(ft.TextSpan(text))
            return spans

        text_lower = text.lower()
        query_lower = query.lower()
        start = 0
        found_any = False
        while True:
            idx = text_lower.find(query_lower, start)
            if idx == -1:
                spans.append(ft.TextSpan(text[start:]))
                break
            found_any = True
            if idx > start:
                spans.append(ft.TextSpan(text[start:idx]))
            spans.append(
                ft.TextSpan(
                    text[idx:idx + len(query)],
                    style=ft.TextStyle(
                        weight=ft.FontWeight.BOLD,
                        bgcolor=ft.Colors.YELLOW_200,
                        color=ft.Colors.BLACK,
                    ),
                )
            )
            start = idx + len(query)
        if not found_any:
            return [ft.TextSpan(text)]
        return spans

    def clean_extracted_text(text):
        """
        خطوط خالی متوالی زیاد را به یک خط خالی کاهش می‌دهد تا فاصله‌های
        بزرگ بین بخش‌های متن (که در فایل اصلی PDF وجود دارد) باعث نشود
        محتوای واقعی خیلی پایین‌تر از دید کاربر بیفتد.
        """
        lines = [ln.rstrip() for ln in text.split("\n")]
        cleaned = []
        blank_run = 0
        for ln in lines:
            if ln.strip() == "":
                blank_run += 1
                if blank_run <= 1:
                    cleaned.append("")
            else:
                blank_run = 0
                cleaned.append(ln)
        return "\n".join(cleaned).strip()

    def extract_pdf_pages_text(pdf_path, start_page, end_page):
        """متن صفحات start_page تا end_page (شماره صفحه از ۱ شروع می‌شود) را استخراج می‌کند."""
        reader = PdfReader(pdf_path)
        total_pages = len(reader.pages)
        start_page = max(1, start_page)
        end_page = min(total_pages, end_page)

        pages_text = []
        for page_num in range(start_page, end_page + 1):
            raw_text = reader.pages[page_num - 1].extract_text() or ""
            pages_text.append((page_num, clean_extracted_text(raw_text)))
        return pages_text

    # هدر خلاصه‌ی نتیجه جستجو (اسم قطعه + بازه صفحات) با طراحی شیک‌تر از status_text ساده
    search_result_header = ft.Container(
        content=ft.Row(
            controls=[
                ft.Icon(ft.Icons.SETTINGS, size=18, color=ft.Colors.BLUE_700),
                ft.Text("", weight=ft.FontWeight.BOLD, size=15, expand=True),
            ],
            spacing=8,
        ),
        padding=ft.padding.symmetric(horizontal=12, vertical=8),
        bgcolor=ft.Colors.BLUE_50,
        border_radius=8,
        visible=False,
    )

    def on_search(e):
        query = search_input.value.strip() if search_input.value else ""
        results_view.controls.clear()
        result_text_controls.clear()
        search_result_header.visible = False

        if not query:
            status_text.value = "عبارت جستجو خالی است."
            page.update()
            return

        excel_path = find_excel_file()
        if not excel_path:
            status_text.value = "هیچ فایل اکسلی برای جستجو پیدا نشد."
            page.update()
            return

        match = find_matching_row(excel_path, query)
        if not match:
            status_text.value = f"نتیجه‌ای برای '{query}' پیدا نشد."
            page.update()
            return

        pdf_filename, start_page, end_page, device_name = match
        pdf_path = os.path.join(docs_dir, pdf_filename)

        if not os.path.exists(pdf_path):
            status_text.value = f"فایل PDF مربوطه ('{pdf_filename}') در پوشه اسناد پیدا نشد."
            page.update()
            return

        try:
            pages_text = extract_pdf_pages_text(pdf_path, start_page, end_page)
        except Exception:
            status_text.value = "خطا در خواندن صفحات فایل PDF."
            page.update()
            return

        status_text.value = ""
        search_result_header.content.controls[1].value = (
            f"قطعه «{device_name}»  •  صفحات {start_page} تا {end_page}"
        )
        search_result_header.visible = True

        for page_num, text in pages_text:
            paragraphs = split_into_paragraphs(text) or ["(متنی برای نمایش یافت نشد)"]

            paragraph_controls = []
            for para in paragraphs:
                spans = build_highlighted_spans(para, query, result_font_size["value"])
                para_text = ft.Text(
                    spans=spans,
                    size=result_font_size["value"],
                    selectable=True,
                    style=ft.TextStyle(height=1.7),
                )
                paragraph_controls.append(para_text)
                result_text_controls.append(para_text)

            page_card = ft.Container(
                content=ft.Column(
                    controls=[
                        ft.Row(
                            controls=[
                                ft.Container(
                                    content=ft.Text(
                                        f"صفحه {page_num}",
                                        size=12,
                                        weight=ft.FontWeight.BOLD,
                                        color=ft.Colors.WHITE,
                                    ),
                                    bgcolor=ft.Colors.BLUE_600,
                                    padding=ft.padding.symmetric(horizontal=10, vertical=4),
                                    border_radius=20,
                                ),
                            ],
                            alignment=ft.MainAxisAlignment.START,
                        ),
                        ft.Divider(height=1, color=ft.Colors.OUTLINE_VARIANT),
                        ft.Column(controls=paragraph_controls, spacing=10),
                    ],
                    spacing=10,
                ),
                padding=14,
                bgcolor=ft.Colors.SURFACE,
                border=ft.border.all(1, ft.Colors.OUTLINE_VARIANT),
                border_radius=12,
                margin=ft.margin.only(bottom=10),
                shadow=ft.BoxShadow(
                    spread_radius=0,
                    blur_radius=6,
                    color=ft.Colors.with_opacity(0.08, ft.Colors.BLACK),
                    offset=ft.Offset(0, 2),
                ),
            )
            results_view.controls.append(page_card)

        page.update()

    def zoom_in(e):
        result_font_size["value"] = min(30, result_font_size["value"] + 2)
        for t in result_text_controls:
            t.size = result_font_size["value"]
        page.update()

    def zoom_out(e):
        result_font_size["value"] = max(10, result_font_size["value"] - 2)
        for t in result_text_controls:
            t.size = result_font_size["value"]
        page.update()

    search_btn = ft.TextButton(
        content=ft.Row([ft.Icon(ft.Icons.SEARCH), ft.Text("جستجو")], spacing=5),
        on_click=on_search
    )

    zoom_out_btn = ft.IconButton(icon=ft.Icons.ZOOM_OUT, on_click=zoom_out, tooltip="کوچک‌نمایی متن")
    zoom_in_btn = ft.IconButton(icon=ft.Icons.ZOOM_IN, on_click=zoom_in, tooltip="بزرگ‌نمایی متن")

    update_dropdown()

    page.add(
        ft.Column(
            controls=[
                ft.Container(height=10),
                ft.Text("سیستم مدیریت اسناد کارخانه", size=18, weight=ft.FontWeight.BOLD),
                ft.Divider(),
                company_dropdown,
                ft.Row([pick_button, process_button], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                ft.Container(height=5),
                status_text,
                ft.Divider(),
                ft.Row([search_input, search_btn]),
                ft.Row(
                    [ft.Text("اندازه متن نتایج:", size=12), zoom_out_btn, zoom_in_btn],
                    alignment=ft.MainAxisAlignment.START,
                ),
                search_result_header,
                results_view,
            ],
            alignment=ft.MainAxisAlignment.START,
            horizontal_alignment=ft.CrossAxisAlignment.STRETCH,
            expand=True
        )
    )

ft.app(target=main)
